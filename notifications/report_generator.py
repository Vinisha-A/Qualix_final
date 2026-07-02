import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.conf import settings
from django.utils import timezone

def generate_excel_report(run):
    """
    Generates an Excel validation report for the given ValidationRun.
    Returns the absolute file path where the report is saved.
    """
    mapping = run.mapping
    results = run.results.select_related('column_mapping').all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Validation Report"

    col_headers = [
        'Validation Rule Name',
        'Source Value',
        'Target Value',
        'Status',
        'Error Message (if failed)'
    ]

    ws.append(col_headers)

    # Style header row
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid") # Deep Blue
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in results:
        status_val = "PASSED" if r.is_match else "FAILED"
        err_msg = r.difference if not r.is_match else ""
        
        rule_name = f"{r.column_mapping.source_column} - {r.operation.replace('_', ' ').title()}"
        
        ws.append([
            rule_name,
            str(r.source_value) if r.source_value is not None else "None",
            str(r.target_value) if r.target_value is not None else "None",
            status_val,
            err_msg
        ])

    # Formatting columns
    for row in range(2, ws.max_row + 1):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            if col in [1, 5]:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            if col == 4:
                if cell.value == "PASSED":
                    cell.font = Font(name="Calibri", size=11, bold=True, color="15803D") # Green
                else:
                    cell.font = Font(name="Calibri", size=11, bold=True, color="B91C1C") # Red

    # Adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 15)

    # Prepare save directory
    reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
    os.makedirs(reports_dir, exist_ok=True)

    # Use run's created_at, with timezone.now() fallback if not saved/None
    val_date = run.created_at.strftime('%Y%m%d_%H%M%S') if run.created_at else timezone.now().strftime('%Y%m%d_%H%M%S')
    filename = f"validation_report_{run.id}_{val_date}.xlsx"
    filepath = os.path.join(reports_dir, filename)
    wb.save(filepath)

    return filepath
